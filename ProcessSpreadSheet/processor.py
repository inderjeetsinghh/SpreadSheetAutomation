from sys import exception

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    try:
        wb = xl.load_workbook(filename)
        sheet = wb['Sheet1']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            corrected_value = cell.value * 0.9
            corrected_price_cell = sheet.cell(row, 4)
            corrected_price_cell.value = corrected_value

        values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
        chart = BarChart()
        chart.add_data(values)

        sheet.add_chart(chart,'e2')
        wb.save(filename)
    except FileNotFoundError:
        print("Entered File Name does not exists in directory")
    except KeyError:
        print("Worksheet does not exists.")
